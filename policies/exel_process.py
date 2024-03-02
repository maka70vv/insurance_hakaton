import openpyxl


def parse(file_path):
    data = {}

    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=1):
        data[row_number] = list(row)

    # Close the workbook
    wb.close()
    return data
